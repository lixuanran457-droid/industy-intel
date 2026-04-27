#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成成人行业情报日报 Excel 文件（修复版）
- 对齐 fetch_news.py / fetch_competitors.py / fetch_keywords.py 的真实输出结构
- 优先使用真实抓取数据
- 仅在无数据时使用最小兜底提示，不再大段写死内容
"""

import json
from datetime import datetime
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def load_json_data(filepath: str):
    """加载JSON数据"""
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"❌ 加载 {filepath} 失败: {e}")
        return None


def safe_str(v: Any) -> str:
    return '' if v is None else str(v)


def dedupe_by_title(items: List[Dict[str, Any]], title_key: str = 'title') -> List[Dict[str, Any]]:
    seen = set()
    out = []
    for x in items:
        t = safe_str(x.get(title_key)).strip()
        if not t:
            continue
        k = t.lower()
        if k in seen:
            continue
        seen.add(k)
        out.append(x)
    return out


def parse_news(news_data: Dict[str, Any]) -> Dict[str, List[Dict[str, Any]]]:
    """将 news_data 标准化为事件与玩法"""
    events: List[Dict[str, Any]] = []
    patterns: List[Dict[str, Any]] = []

    if not news_data:
        return {'events': events, 'patterns': patterns}

    date_str = safe_str(news_data.get('date', datetime.now().strftime('%Y-%m-%d')))

    # 中文论坛帖子
    for item in news_data.get('chinese_caoliu_posts', []) or []:
        title = safe_str(item.get('title'))
        events.append({
            'title': title,
            'link': safe_str(item.get('url', '')),
            'source': safe_str(item.get('source', '草榴')),
            'date': date_str
        })

        if any(k in title for k in ['新站', '平台', 'APP', '运营', '变现', '推广', '直播']):
            patterns.append({
                'title': title,
                'source': safe_str(item.get('source', '草榴')),
                'desc': f"论坛信号：{title}",
                'feasibility': '中'
            })

    # 英文新闻
    for src_key, src_name in [('xbiz', 'XBIZ'), ('avn', 'AVN'), ('onlyfans_blog', 'OnlyFans Blog')]:
        for item in news_data.get(src_key, []) or []:
            title = safe_str(item.get('title'))
            events.append({
                'title': title,
                'link': safe_str(item.get('url', '')),
                'source': src_name,
                'date': date_str
            })

            if any(k in title.lower() for k in ['launch', 'feature', 'creator', 'subscription', 'program', 'tool', 'policy']):
                patterns.append({
                    'title': title,
                    'source': src_name,
                    'desc': safe_str(item.get('url', '')),
                    'feasibility': '中'
                })

    events = dedupe_by_title(events, 'title')
    patterns = dedupe_by_title(patterns, 'title')
    return {'events': events[:30], 'patterns': patterns[:20]}


def parse_competitors(competitors_data: Dict[str, Any]) -> Dict[str, List[Dict[str, Any]]]:
    """标准化竞品数据"""
    platforms: List[Dict[str, Any]] = []
    forum_signals: List[Dict[str, Any]] = []

    if not competitors_data:
        return {'platforms': platforms, 'forum_signals': forum_signals}

    for p in competitors_data.get('known_competitors', []) or []:
        platforms.append({
            'name': safe_str(p.get('name')),
            'url': safe_str(p.get('url')),
            'type': safe_str(p.get('type', '待分析')),
            'watch': safe_str(p.get('watch', '待评估')),
            'tier': safe_str(p.get('tier', '')),
        })

    # 从中文论坛信号里保留标题供参考
    for x in competitors_data.get('chinese_forum_posts', []) or []:
        title = safe_str(x.get('title'))
        if title:
            forum_signals.append({
                'title': title,
                'url': safe_str(x.get('url', '')),
                'source': safe_str(x.get('forum', '草榴'))
            })

    platforms = dedupe_by_title(platforms, 'name')
    forum_signals = dedupe_by_title(forum_signals, 'title')
    return {'platforms': platforms[:20], 'forum_signals': forum_signals[:10]}


def parse_monetization(news_data: Dict[str, Any], competitors_data: Dict[str, Any]) -> List[Dict[str, str]]:
    """从真实新闻与竞品关注点提炼商业化模式"""
    rows: List[Dict[str, str]] = []

    # 从 XBIZ/AVN 标题提炼
    for src_key, src_name in [('xbiz', 'XBIZ'), ('avn', 'AVN')]:
        for item in (news_data or {}).get(src_key, []) or []:
            title = safe_str(item.get('title'))
            title_l = title.lower()
            if any(k in title_l for k in ['creator', 'subscription', 'program', 'revenue', 'affiliate', 'studio', 'live']):
                rows.append({
                    'mode': f"{src_name}动态：{title[:48]}",
                    'desc': '行业新闻中出现创作者/订阅/营收相关信号，建议跟踪对应产品动作',
                    'source': safe_str(item.get('url', '')) or src_name
                })

    # 从竞品 watch 字段提炼
    for p in (competitors_data or {}).get('known_competitors', []) or []:
        watch = safe_str(p.get('watch'))
        if watch:
            rows.append({
                'mode': f"{safe_str(p.get('name'))} 关注点",
                'desc': watch,
                'source': safe_str(p.get('url', ''))
            })

    # 去重
    uniq = []
    seen = set()
    for r in rows:
        k = (r['mode'] + '|' + r['source']).lower()
        if k in seen:
            continue
        seen.add(k)
        uniq.append(r)

    if not uniq:
        uniq = [{
            'mode': '暂无足够商业化信号',
            'desc': '本次抓取未命中可结构化商业化条目，建议补充中文导航站/论坛手工记录后再生成',
            'source': '系统提示'
        }]

    return uniq[:20]


def build_actions(keywords_data: Dict[str, Any], competitors_data: Dict[str, Any], news_data: Dict[str, Any]) -> List[List[str]]:
    """基于真实数据生成行动项"""
    actions: List[List[str]] = []

    # 关键词驱动
    top_kw = (keywords_data or {}).get('top_keywords', []) or []
    for kw in top_kw[:3]:
        k = safe_str(kw.get('keyword'))
        opp = safe_str(kw.get('opportunity', '中'))
        prio = '高' if opp == '高' else '中'
        actions.append([
            f"围绕关键词「{k}」抽样拆解 TOP3 页面/频道内容结构",
            prio,
            f"验证 {k} 的真实需求强度，形成选题或投放线索"
        ])

    # 新兴竞品驱动
    for p in (competitors_data or {}).get('known_competitors', []) or []:
        if safe_str(p.get('tier')) == '新兴':
            name = safe_str(p.get('name'))
            watch = safe_str(p.get('watch', '持续跟踪'))
            actions.append([
                f"跟踪新兴竞品 {name} 的近7天动态",
                '高',
                f"建立竞品时间线：{watch}"
            ])

    # 新闻驱动
    xbiz_n = len((news_data or {}).get('xbiz', []) or [])
    avn_n = len((news_data or {}).get('avn', []) or [])
    actions.append([
        f"复盘英文行业源（XBIZ={xbiz_n}, AVN={avn_n}）并筛选可本地化玩法",
        '中',
        '补足中文源不稳定时的信息覆盖'
    ])

    # 去重 + 截断
    out = []
    seen = set()
    for a in actions:
        k = a[0]
        if k in seen:
            continue
        seen.add(k)
        out.append(a)

    if not out:
        out = [["暂无可执行项，请先执行抓取脚本", "中", "补齐数据后自动生成行动建议"]]

    return out[:10]


def create_excel_report(news_data, competitors_data, keywords_data, output_path):
    wb = Workbook()
    wb.remove(wb.active)

    date_str = datetime.now().strftime('%Y-%m-%d')

    # 样式定义
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    section_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    section_font = Font(bold=True, size=11)

    parsed_news = parse_news(news_data or {})
    parsed_comp = parse_competitors(competitors_data or {})
    monetization_rows = parse_monetization(news_data or {}, competitors_data or {})
    actions = build_actions(keywords_data or {}, competitors_data or {}, news_data or {})

    # ===== Sheet 1: 今日大事件 =====
    ws1 = wb.create_sheet('今日大事件', 0)
    ws1.append([f'📰 成人行业情报日报 - {date_str}'])
    ws1.merge_cells('A1:D1')
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A1'].alignment = Alignment(horizontal='center')

    ws1.append([])
    ws1.append(['一、今日大事件'])
    ws1['A3'].font = section_font
    ws1['A3'].fill = section_fill
    ws1.append(['标题', '链接', '来源', '日期'])
    for cell in ws1[4]:
        cell.font = header_font
        cell.fill = header_fill

    for event in parsed_news['events']:
        ws1.append([
            safe_str(event.get('title')),
            safe_str(event.get('link')),
            safe_str(event.get('source')),
            safe_str(event.get('date', date_str))
        ])

    if not parsed_news['events']:
        ws1.append(['⚠ 未抓到可用大事件，请先执行 fetch_news.py', '', '系统提示', date_str])

    ws1.column_dimensions['A'].width = 58
    ws1.column_dimensions['B'].width = 62
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 12

    # ===== Sheet 2: 最新玩法 =====
    ws2 = wb.create_sheet('最新玩法', 1)
    ws2.append(['二、最新玩法'])
    ws2['A1'].font = section_font
    ws2['A1'].fill = section_fill
    ws2.append(['玩法', '来源', '描述', '可落地性'])
    for cell in ws2[2]:
        cell.font = header_font
        cell.fill = header_fill

    for p in parsed_news['patterns']:
        ws2.append([
            safe_str(p.get('title')),
            safe_str(p.get('source')),
            safe_str(p.get('desc')),
            safe_str(p.get('feasibility', '中')),
        ])

    if not parsed_news['patterns']:
        ws2.append(['⚠ 暂无结构化玩法信号', '系统提示', '可检查中文论坛/竞品动态后重跑', '低'])

    ws2.column_dimensions['A'].width = 52
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 68
    ws2.column_dimensions['D'].width = 12

    # ===== Sheet 3: 竞品黑马 =====
    ws3 = wb.create_sheet('竞品黑马', 2)
    ws3.append(['三、竞品黑马'])
    ws3['A1'].font = section_font
    ws3['A1'].fill = section_fill
    ws3.append(['平台', '网址', '类型', '值得关注'])
    for cell in ws3[2]:
        cell.font = header_font
        cell.fill = header_fill

    for p in parsed_comp['platforms']:
        name = safe_str(p.get('name'))
        tier = safe_str(p.get('tier'))
        display_name = f"{name} [{tier}]" if tier else name
        ws3.append([
            display_name,
            safe_str(p.get('url')),
            safe_str(p.get('type', '待分析')),
            safe_str(p.get('watch', '待评估'))
        ])

    # 附加中文论坛竞品信号
    if parsed_comp['forum_signals']:
        ws3.append(['--- 论坛信号 ---', '', '', ''])
        for s in parsed_comp['forum_signals'][:5]:
            ws3.append([
                safe_str(s.get('title'))[:60],
                safe_str(s.get('url')),
                safe_str(s.get('source')),
                '来自论坛讨论，建议人工核验'
            ])

    if not parsed_comp['platforms']:
        ws3.append(['⚠ 暂无竞品数据', '', '', '请先执行 fetch_competitors.py'])

    ws3.column_dimensions['A'].width = 44
    ws3.column_dimensions['B'].width = 62
    ws3.column_dimensions['C'].width = 18
    ws3.column_dimensions['D'].width = 38

    # ===== Sheet 4: 商业化新模式 =====
    ws4 = wb.create_sheet('商业化新模式', 3)
    ws4.append(['四、商业化新模式'])
    ws4['A1'].font = section_font
    ws4['A1'].fill = section_fill
    ws4.append(['模式', '描述', '来源'])
    for cell in ws4[2]:
        cell.font = header_font
        cell.fill = header_fill

    for r in monetization_rows:
        ws4.append([
            safe_str(r.get('mode')),
            safe_str(r.get('desc')),
            safe_str(r.get('source')),
        ])

    ws4.column_dimensions['A'].width = 44
    ws4.column_dimensions['B'].width = 86
    ws4.column_dimensions['C'].width = 62

    # ===== Sheet 5: 关键词日榜 =====
    ws5 = wb.create_sheet('关键词日榜', 4)
    ws5.append(['五、关键词日榜 TOP 10'])
    ws5['A1'].font = section_font
    ws5['A1'].fill = section_fill
    ws5.append(['排名', '关键词', '热度', '趋势', '商机'])
    for cell in ws5[2]:
        cell.font = header_font
        cell.fill = header_fill

    top_keywords = (keywords_data or {}).get('top_keywords', []) or []
    for kw in top_keywords[:10]:
        ws5.append([
            kw.get('rank', ''),
            kw.get('keyword', ''),
            kw.get('score', ''),
            kw.get('trend', ''),
            kw.get('opportunity', ''),
        ])

    if not top_keywords:
        ws5.append(['', '⚠ 无关键词数据', '', '', '请先执行 fetch_keywords.py'])

    ws5.column_dimensions['A'].width = 8
    ws5.column_dimensions['B'].width = 28
    ws5.column_dimensions['C'].width = 10
    ws5.column_dimensions['D'].width = 10
    ws5.column_dimensions['E'].width = 12

    # ===== Sheet 6: 可行动项 =====
    ws6 = wb.create_sheet('可行动项', 5)
    ws6.append(['六、可行动项'])
    ws6['A1'].font = section_font
    ws6['A1'].fill = section_fill
    ws6.append(['行动', '优先级', '预期效果'])
    for cell in ws6[2]:
        cell.font = header_font
        cell.fill = header_fill

    for action in actions:
        ws6.append(action)

    ws6.column_dimensions['A'].width = 76
    ws6.column_dimensions['B'].width = 10
    ws6.column_dimensions['C'].width = 44

    wb.save(output_path)
    print(f'✅ Excel 日报已生成: {output_path}')
    return output_path


def main():
    news_data = load_json_data('/tmp/news_data.json')
    competitors_data = load_json_data('/tmp/competitors_data.json')
    keywords_data = load_json_data('/tmp/keywords_data.json')

    date_str = datetime.now().strftime('%Y%m%d')
    output_path = f'/tmp/industry_intel_report_{date_str}.xlsx'
    create_excel_report(news_data, competitors_data, keywords_data, output_path)

    print(f'\n📊 报告路径: {output_path}')
    return output_path


if __name__ == '__main__':
    main()

